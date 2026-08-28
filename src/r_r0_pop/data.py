from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PREIMAGINAL_STAGES = ("E", "L1", "L2", "L3", "P")
BASER_POOLED_FEMALE_ADULTS = 152
BASER_POOLED_MALE_ADULTS = 105
BASER_POOLED_ADULT_FEMALE_FRACTION_UNROUNDED = BASER_POOLED_FEMALE_ADULTS / (
    BASER_POOLED_FEMALE_ADULTS + BASER_POOLED_MALE_ADULTS
)
BASER_POOLED_ADULT_FEMALE_FRACTION = BASER_POOLED_ADULT_FEMALE_FRACTION_UNROUNDED


@dataclass(frozen=True)
class BaserPaths:
    """Paths to the Baser et al. workbooks."""

    life_tables: Path
    fertility: Path

    @classmethod
    def from_data_dir(cls, data_dir: Path | str) -> "BaserPaths":
        data_dir = Path(data_dir)
        return cls(
            life_tables=data_dir / "LifeTablesDataset.xlsx",
            fertility=data_dir / "Fertility.xlsx",
        )


@dataclass(frozen=True)
class BaserProcessedPaths:
    """Paths to tidy CSV files derived from the Baser workbooks."""

    development: Path
    adult_survival: Path
    fertility: Path

    @classmethod
    def from_processed_dir(cls, processed_dir: Path | str) -> "BaserProcessedPaths":
        processed_dir = Path(processed_dir)
        return cls(
            development=processed_dir / "development.csv",
            adult_survival=processed_dir / "adult_survival.csv",
            fertility=processed_dir / "fertility.csv",
        )

    def exists(self) -> bool:
        return (
            self.development.exists()
            and self.adult_survival.exists()
            and self.fertility.exists()
        )


def parse_temperature(label: object) -> float | None:
    """Return the first temperature value from labels such as '24 C' or '24C'."""

    if label is None:
        return None
    match = re.search(r"(-?\d+(?:\.\d+)?)\s*C?", str(label))
    return float(match.group(1)) if match else None


def load_baser_life_history(path: Path | str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load individual stage durations from Baser's life-history workbook.

    Returns
    -------
    development:
        One row per specimen and temperature with durations for successful
        preimaginal stage transitions.
    adult_survival:
        One row per specimen and temperature with adult sex-specific longevity
        and death-stage durations.
    """

    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb["Individual-LifeHistory"]

    blocks: list[tuple[int, float]] = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(20, col).value
        if isinstance(value, str) and "Development - Condition" in value:
            temp = parse_temperature(value)
            if temp is not None:
                blocks.append((col, temp))

    development_rows: list[dict[str, float | int]] = []
    adult_rows: list[dict[str, float | int]] = []

    for start_col, temp in blocks:
        header = {
            ws.cell(21, col).value: col
            for col in range(start_col, min(start_col + 13, ws.max_column + 1))
        }
        specimen_col = header["Specimen"]

        for row in range(22, 1022):
            specimen = ws.cell(row, specimen_col).value
            if not isinstance(specimen, (int, float)):
                continue
            record: dict[str, float | int] = {
                "temperature": temp,
                "specimen": int(specimen),
            }
            for stage in PREIMAGINAL_STAGES:
                value = ws.cell(row, header[stage]).value
                record[stage] = _numeric_or_zero(value)
            development_rows.append(record)

        adult_header = {
            ws.cell(1026, col).value: col
            for col in range(start_col, min(start_col + 13, ws.max_column + 1))
        }
        adult_specimen_col = adult_header["Specimen"]
        for row in range(1027, 2027):
            specimen = ws.cell(row, adult_specimen_col).value
            if not isinstance(specimen, (int, float)):
                continue
            record = {
                "temperature": temp,
                "specimen": int(specimen),
                "AM": _numeric_or_zero(ws.cell(row, adult_header["AM"]).value),
                "AF": _numeric_or_zero(ws.cell(row, adult_header["AF"]).value),
            }
            for stage in PREIMAGINAL_STAGES:
                record[f"death_{stage}"] = _numeric_or_zero(
                    ws.cell(row, adult_header[stage]).value
                )
            adult_rows.append(record)

    return pd.DataFrame(development_rows), pd.DataFrame(adult_rows)


def load_baser_fertility(path: Path | str) -> pd.DataFrame:
    """Load individual daily egg records from Baser's fertility workbook.

    Missing egg counts are kept as missing values. In the Baser sheet, daily
    averages exclude those missing values, which likely represent females no
    longer observed/alive. Downstream code treats non-missing records as live
    female-days.
    """

    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb["Dataset"]
    rows: list[dict[str, float | int | str | None]] = []

    row = 1
    while row <= ws.max_row:
        temp = parse_temperature(ws.cell(row, 1).value)
        if temp is None:
            row += 1
            continue

        preovip_row = row + 2
        preovip = {
            female_idx: _optional_numeric(ws.cell(preovip_row, female_idx + 1).value)
            for female_idx in range(1, 11)
        }

        day_header_row = None
        for candidate in range(row + 1, min(row + 8, ws.max_row + 1)):
            if ws.cell(candidate, 1).value == "Day":
                day_header_row = candidate
                break
        if day_header_row is None:
            row += 1
            continue

        current = day_header_row + 1
        while current <= ws.max_row and isinstance(
            ws.cell(current, 1).value, (int, float)
        ):
            adult_day = int(ws.cell(current, 1).value)
            for female_idx in range(1, 11):
                eggs = _optional_numeric(ws.cell(current, female_idx + 1).value)
                rows.append(
                    {
                        "temperature": temp,
                        "female": f"F{female_idx}",
                        "female_index": female_idx,
                        "preoviposition_days": preovip[female_idx],
                        "adult_day": adult_day,
                        "eggs": eggs,
                    }
                )
            current += 1

        row = current + 1

    return pd.DataFrame(rows)


def write_baser_processed_data(
    raw_paths: BaserPaths, processed_dir: Path | str
) -> BaserProcessedPaths:
    """Convert Baser Excel workbooks into tidy CSV files."""

    processed_paths = BaserProcessedPaths.from_processed_dir(processed_dir)
    processed_paths.development.parent.mkdir(parents=True, exist_ok=True)

    development, adult_survival = load_baser_life_history(raw_paths.life_tables)
    fertility = load_baser_fertility(raw_paths.fertility)

    development.to_csv(processed_paths.development, index=False)
    adult_survival.to_csv(processed_paths.adult_survival, index=False)
    fertility.to_csv(processed_paths.fertility, index=False)

    return processed_paths


def load_baser_processed_data(
    processed_dir: Path | str,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load tidy CSV files produced by `write_baser_processed_data`."""

    processed_paths = BaserProcessedPaths.from_processed_dir(processed_dir)
    if not processed_paths.exists():
        missing = [
            str(path)
            for path in (
                processed_paths.development,
                processed_paths.adult_survival,
                processed_paths.fertility,
            )
            if not path.exists()
        ]
        raise FileNotFoundError(f"Missing processed Baser data: {', '.join(missing)}")

    development = pd.read_csv(processed_paths.development)
    adult_survival = pd.read_csv(processed_paths.adult_survival)
    fertility = pd.read_csv(processed_paths.fertility)
    return development, adult_survival, fertility


def load_or_create_baser_processed_data(
    raw_paths: BaserPaths,
    processed_dir: Path | str,
    *,
    rebuild: bool = False,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load processed Baser CSVs, creating them from Excel when needed."""

    processed_paths = BaserProcessedPaths.from_processed_dir(processed_dir)
    if rebuild or not processed_paths.exists():
        write_baser_processed_data(raw_paths, processed_dir)
    return load_baser_processed_data(processed_dir)


def female_preadult_summary(
    development: pd.DataFrame, adult_survival: pd.DataFrame, cohort_size: int = 50
) -> pd.DataFrame:
    """Summarize juvenile survival and preimaginal duration for female schedules.

    Juvenile survival uses emergence of either sex and is interpreted as the
    survival probability of an initial female egg under sex-independent
    juvenile survival. Timing uses females because the reproduction schedule is
    measured for adult females.
    """

    female_adults = adult_survival.loc[
        adult_survival["AF"] > 0, ["temperature", "specimen"]
    ]
    merged = female_adults.merge(
        development, on=["temperature", "specimen"], how="left"
    )
    merged["preadult_days"] = merged[list(PREIMAGINAL_STAGES)].sum(axis=1)
    timing = (
        merged.groupby("temperature", as_index=False)
        .agg(
            female_adults=("specimen", "count"),
            mean_female_preadult_days=("preadult_days", "mean"),
        )
        .sort_values("temperature")
    )
    emerged = adult_survival.assign(
        emerged_adult=(adult_survival["AF"] > 0) | (adult_survival["AM"] > 0)
    )
    survival = (
        emerged.groupby("temperature", as_index=False)
        .agg(adult_count=("emerged_adult", "sum"))
        .sort_values("temperature")
    )
    survival["p_survive_to_adult"] = survival["adult_count"] / cohort_size
    summary = timing.merge(survival, on="temperature", how="left")
    return summary


def female_preadult_distribution(
    development: pd.DataFrame,
    adult_survival: pd.DataFrame,
    cohort_size: int = 50,
) -> pd.DataFrame:
    """Return the empirical female preadult-duration distribution.

    Fertility females cannot be linked to specimens in the life-table cohort.
    The direct demographic calculation therefore treats the observed preadult
    durations of emerged females as an empirical timing distribution and
    combines it independently with the observed adult reproduction schedules.
    Each row receives equal weight within temperature. Juvenile survival is
    estimated from emergence of either sex, as in ``female_preadult_summary``.
    """

    female_adults = adult_survival.loc[
        adult_survival["AF"] > 0, ["temperature", "specimen"]
    ]
    timing = female_adults.merge(
        development, on=["temperature", "specimen"], how="left"
    )
    timing["preadult_days"] = timing[list(PREIMAGINAL_STAGES)].sum(axis=1)
    timing = timing.loc[timing["preadult_days"] > 0].copy()
    timing["female_adults"] = timing.groupby("temperature")["specimen"].transform(
        "size"
    )
    timing["preadult_weight"] = 1.0 / timing["female_adults"]

    emerged = adult_survival.assign(
        emerged_adult=(adult_survival["AF"] > 0) | (adult_survival["AM"] > 0)
    )
    survival = (
        emerged.groupby("temperature", as_index=False)
        .agg(adult_count=("emerged_adult", "sum"))
        .sort_values("temperature")
    )
    survival["p_survive_to_adult"] = survival["adult_count"] / cohort_size
    return (
        timing[
            [
                "temperature",
                "specimen",
                "preadult_days",
                "female_adults",
                "preadult_weight",
            ]
        ]
        .merge(survival, on="temperature", how="left")
        .sort_values(["temperature", "specimen"])
        .reset_index(drop=True)
    )


def pooled_adult_female_fraction(adult_survival: pd.DataFrame) -> float:
    """Return the female fraction among all sexed adults pooled over temperature."""

    female = adult_survival["AF"] > 0
    male = adult_survival["AM"] > 0
    if bool((female & male).any()):
        raise ValueError("Adult records cannot be both female and male.")
    adult_count = int((female | male).sum())
    if adult_count == 0:
        raise ValueError(
            "Cannot estimate the female fraction because no sexed adults were found."
        )
    return float(female.sum() / adult_count)


def _numeric_or_zero(value: object) -> float:
    return float(value) if isinstance(value, (int, float)) else 0.0


def _optional_numeric(value: object) -> float | None:
    return float(value) if isinstance(value, (int, float)) else None
